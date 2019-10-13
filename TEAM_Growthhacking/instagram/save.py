import time
from selenium import webdriver
from openpyxl import load_workbook

# https://www.instagram.com/p/B3JjzSPBnGF/
# https://www.instagram.com/p/B3HviHRIVDn/
# 위의 두 url만 보면 중간 부문만 달라짐

URL = 'https://www.instagram.com/explore/tags/{}'
DRIVER_DIR = 'C:/Users/dkdud/Desktop/fastcampus/chromedriver'
SAVE_DIR = 'test.xlsx'

def instagram_scrap(keyword):
    try:
        driver = webdriver.Chrome(DRIVER_DIR)
        driver.implicitly_wait(10)
        driver.get(URL.format(keyword))

        new_links = driver.find_elements_by_css_selector('div.v1Nh3 > a')
        links = [i.get_attribute('href') for i in new_links]

        result = []

        for link in links:
            driver.get(link)
            for li in driver.find_elements_by_class_name('C4VMK'):
                user = li.find_elements_by_tag_name('a')[0].text #작성자
                # text - 요소 내부의 글자
                reply = li.find_elements_by_tag_name('span')[0].text #댓글, 해시태그
                result.append("({}) {}".format(user, reply))
                # print("({}) {}".format(user, reply))
    
    except Exception as e:
        print(e)
    finally:
        driver.close()

def save_excel(result):
    try: 
        wb = load_workbook(SAVE_DIR)
        ws = wb.create_sheet(title='instagram')
        for idx, re in enumerate(result):
            ws['A' + str(idx + 1)] = re[0]
        wb.save(SAVE_DIR)
    finally:
        wb.close()

# def get_excel():
    # wb = load_workbook(SAVE_DIR, read_only=True)
    # ws = wb['instagram']
    # for low in ws.rows:
        # for cell in row:
            # print(cell.value)

if __name__ == "__main__":
    keyword = input('keyword(tag)')
    instagram_scrap(str(keyword))
    